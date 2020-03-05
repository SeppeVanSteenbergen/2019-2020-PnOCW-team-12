import math
import random
import time

import openpyxl
import cv2
import glob
import os
import score


def cut_image(file):
    image = cv2.imread(file)
    row, column, channels = image.shape
    top_left = [int(row * (2 / 5)), int(column * (2 / 5))]
    bottom_right = [int(row * (3 / 5)), int(column * (3 / 5))]
    return image[top_left[0]:bottom_right[0], top_left[1]:bottom_right[1], :]


def addToExcell(shet, iid, blocknb, ksize, orColor, Env, Light, Brightness, nbSc, GBsc, GBt, MBs, MBt, MeanBs, MeanBt, FDs, FDt):
    shet.cell(iid + 2, 1, iid)
    shet.cell(iid + 2, 2, blocknb)
    shet.cell(iid + 2, 3, ksize)
    shet.cell(iid + 2, 4, orColor)
    shet.cell(iid + 2, 5, Env)
    shet.cell(iid + 2, 6, Light)
    shet.cell(iid + 2, 7, Brightness)
    shet.cell(iid + 2, 8, nbSc)
    shet.cell(iid + 2, 9, GBsc)
    shet.cell(iid + 2, 10, GBt)
    shet.cell(iid + 2, 11, MBs)
    shet.cell(iid + 2, 12, MBt)
    shet.cell(iid + 2, 13, MeanBs)
    shet.cell(iid + 2, 14, MeanBt)
    shet.cell(iid + 2, 15, FDs)
    shet.cell(iid + 2, 16, FDt)



def setupSheet(sheet, ksize):
    sheet.title = "results " + str(ksize)
    sheet.cell(1, 1, "Id")
    sheet.cell(1, 2, "nbBlocks")
    sheet.cell(1, 3, "kSize")
    sheet.cell(1, 4, "Original color")
    sheet.cell(1, 5, "Environment")
    sheet.cell(1, 6, "Light")
    sheet.cell(1, 7, "Brightness")
    sheet.cell(1, 8, "no blur score")
    sheet.cell(1, 9, "gaussian blur score")
    sheet.cell(1, 10, "gaussian blur time")
    sheet.cell(1, 11, "median blur score")
    sheet.cell(1, 12, "median blur time")
    sheet.cell(1, 13, "mean blur score")
    sheet.cell(1, 14, "mean blur time")
    sheet.cell(1, 15, "fastNM denoising score")
    sheet.cell(1, 16, "fastNM denoising time")


path = os.getcwd()
files = [file for file in glob.glob(os.path.join(path, "**/*.jpg"), recursive=True)]


def calcRandomPixels(width, height, nbBlocks):
    nbBlockSplit = int(math.sqrt(nbBlocks))
    blockWidth = width // nbBlockSplit
    blockHeight = height // nbBlockSplit
    randomPixels = []

    for j in range(int(blockWidth * blockHeight * 0.1)):
        x = random.randint(0, blockWidth - 1)
        y = random.randint(0, blockHeight - 1)
        randomPixels.append([x, y])

    return randomPixels


for ksize in range(3, 13, 2):
    print("--------------start--------------")
    print("--------------" + str(ksize) + "--------------")
    wb = openpyxl.Workbook()
    sheet = wb.active
    setupSheet(sheet, ksize)
    nbBlocks = 4000;
    for i in range(len(files)):
        file = files[i]
        properties = file.split("\\")[-1].split(".")[0].split("_")
        environment, light, brightness, orgColor = properties[0], properties[1], properties[2], properties[3]
        img = cut_image(file)
        height, width = cv2.split(img)[0].shape

        randomPixels = calcRandomPixels(width, height, nbBlocks)
        #no blur
        b, g, r = cv2.split(img)
        noBlurScr = score.scoreImg(r, g, b, orgColor, nbBlocks, randomPixels)

        #Gaussian blur
        start = time.time_ns()
        output = cv2.GaussianBlur(img, (ksize, ksize), 0)
        stop = time.time_ns()

        b, g, r = cv2.split(output)
        gbScr = score.scoreImg(r, g, b, orgColor, nbBlocks, randomPixels)
        gbTime = stop - start

        #Median blur
        start = time.time_ns()
        output = cv2.medianBlur(img, ksize)
        stop = time.time_ns()

        b, g, r = cv2.split(output)
        medianScr = score.scoreImg(r, g, b, orgColor, nbBlocks, randomPixels)
        medianTime = stop - start

        #Mean blur
        start = time.time_ns()
        output = cv2.blur(img, (ksize, ksize))
        stop = time.time_ns()

        b, g, r = cv2.split(output)
        meanScr = score.scoreImg(r, g, b, orgColor, nbBlocks, randomPixels)
        meanTime = stop - start

        #fast denoising shit
        start = time.time_ns()
        cv2.fastNlMeansDenoisingColored(img, output)
        stop = time.time_ns()

        b, g, r = cv2.split(output)
        fastScr = score.scoreImg(r, g, b, orgColor, nbBlocks, randomPixels)
        fastTime = stop - start

        print("_".join(properties) + " done")
        addToExcell(sheet, i, nbBlocks, ksize, orgColor, environment, light, brightness,
                    noBlurScr, gbScr, gbTime, medianScr, medianTime, meanScr, meanTime,
                    fastScr, fastTime)

    file = os.path.join(os.getcwd(), "result_" + str(nbBlocks) + "_" + str(int(time.time())) + ".xls")
    wb.save(file)
    print("--------------end--------------")

